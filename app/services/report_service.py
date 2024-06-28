from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from io import BytesIO
from app.model import Order, Status
import xml.etree.ElementTree as ET
from app.extensions import db
from flask import jsonify
from .utils.validation import validate_order_data
import h5py


def generate_xlsx_report():
    colors = {
        Status.NEW.value: "2192FF",
        Status.IN_PROGRESS.value: "FDFF00",
        Status.COMPLETED.value: "38E54D",
    }

    thin_border = Border(
        left=Side(border_style=BORDER_THIN),
        right=Side(border_style=BORDER_THIN),
        top=Side(border_style=BORDER_THIN),
        bottom=Side(border_style=BORDER_THIN),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.append(["ID", "Name", "Description", "Creation Date", "Status"])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = thin_border

    orders = Order.query.order_by(Order.status, Order.id).all()
    for order in orders:
        row = [
            order.id,
            order.name,
            order.description,
            order.creation_date.strftime("%Y-%m-%d %H:%M:%S"),
            order.status.value,
        ]
        ws.append(row)
        fill = PatternFill(
            start_color=colors[order.status.value],
            end_color=colors[order.status.value],
            fill_type="solid",
        )
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.border = thin_border

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    report = BytesIO()
    wb.save(report)
    report.seek(0)

    return report


def export_to_xml():
    orders = Order.query.order_by(Order.id).all()

    root = ET.Element("Orders")

    for order in orders:
        sub_element = ET.SubElement(root, "Order")
        id_element = ET.SubElement(sub_element, "ID")
        id_element.text = str(order.id)
        name_element = ET.SubElement(sub_element, "Name")
        name_element.text = order.name
        description_element = ET.SubElement(sub_element, "Description")
        description_element.text = order.description
        creation_date_element = ET.SubElement(sub_element, "CreationDate")
        creation_date_element.text = order.creation_date.strftime("%Y-%m-%d %H:%M:%S")
        status_element = ET.SubElement(sub_element, "Status")
        status_element.text = order.status.value

    tree = ET.ElementTree(root)
    xml = BytesIO()
    tree.write(xml, encoding="utf-8", xml_declaration=True)
    xml.seek(0)

    return xml


def import_from_xml(xml):
    try:
        root = ET.fromstring(xml)
    except ET.ParseError:
        return jsonify({"error": "Invalid format."}), 400

    for order_element in root.findall("Order"):
        try:
            id_element = order_element.find("ID")
            name_element = order_element.find("Name")
            description_element = order_element.find("Description")
            creation_date_element = order_element.find("CreationDate")
            status_element = order_element.find("Status")

            if (
                id_element is None
                or name_element is None
                or description_element is None
                or creation_date_element is None
                or status_element is None
            ):
                return (
                    jsonify({"error": "Missing one of required elements in Order."}),
                    400,
                )

            data = {
                "id": id_element.text,
                "name": name_element.text,
                "description": description_element.text,
                "creation_date": creation_date_element.text,
                "status": status_element.text,
            }

            errors = validate_order_data(data)
            if errors["errors"]:
                return jsonify(errors), 400

            __update_or_create_order(data)

        except Exception as e:
            return jsonify({"error": f"Invalid data: {e}"}), 400

    db.session.commit()

    return jsonify({"message": "Orders were imported successfully from XML."}), 200


def export_to_hdf5():
    orders = Order.query.all()

    hdf5 = BytesIO()

    with h5py.File(hdf5, "w") as file:
        for order in orders:
            order_group = file.create_group(f"order_{order.id}")
            order_group.attrs["id"] = order.id
            order_group.attrs["name"] = order.name
            order_group.attrs["description"] = order.description
            order_group.attrs["creation_date"] = order.creation_date.strftime(
                "%Y-%m-%d %H:%M:%S"
            )
            order_group.attrs["status"] = order.status.value

    hdf5.seek(0)

    return hdf5


def import_from_hdf5(hdf5):
    try:
        with h5py.File(hdf5, "r") as file:
            for key in file.keys():
                order_group = file[key]
                data = {
                    "id": order_group.attrs["id"],
                    "name": order_group.attrs["name"],
                    "description": order_group.attrs["description"],
                    "creation_date": order_group.attrs["creation_date"],
                    "status": order_group.attrs["status"],
                }

                errors = validate_order_data(data)
                if errors["errors"]:
                    return jsonify(errors), 400

                __update_or_create_order(data)

            db.session.commit()

            return (
                jsonify(
                    {"message": "Orders were imported successfully from HDF5 file."}
                ),
                200,
            )
    except Exception as e:
        return jsonify({"error": f"Failed to import from HDF5: {str(e)}"}), 400


def __update_or_create_order(data):
    existing_order = db.session.get(Order, int(data["id"]))
    if existing_order:
        existing_order.name = data["name"]
        existing_order.description = data["description"]
        existing_order.status = data["status"]
    else:
        order = Order(
            name=data["name"],
            description=data["description"],
            creation_date=data["creation_date"],
            status=data["status"],
        )
        db.session.add(order)
