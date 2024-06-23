from io import BytesIO
from openpyxl import load_workbook
from app.model import Status, Order


def test_generate_xlsx_report(client, init_orders):
    response = client.get("/api/orders/xlsx-report")
    assert response.status_code == 200
    assert (
        response.headers["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    xlsx_report = BytesIO(response.data)
    wb = load_workbook(xlsx_report)
    ws = wb.active

    assert ws.title == "Orders"

    required_headers = ["ID", "Name", "Description", "Creation Date", "Status"]
    for i, cell in enumerate(ws[1]):
        assert cell.value == required_headers[i]

    orders = Order.query.order_by(Order.status, Order.id).all()

    row_count = len(
        [row for row in ws if not all([cell.value == None for cell in row])]
    )
    assert row_count == len(orders) + 1

    colors = {
        Status.NEW.value: "2192FF",
        Status.IN_PROGRESS.value: "FDFF00",
        Status.COMPLETED.value: "38E54D",
    }

    for i, order in enumerate(orders, start=2):
        row = ws[i]
        assert row[0].value == order.id
        assert row[1].value == order.name
        assert row[2].value == order.description
        assert row[3].value == order.creation_date.strftime("%Y-%m-%d %H:%M:%S")
        assert row[4].value == order.status.value

        for cell in row:
            assert cell.fill.start_color.rgb == "00" + colors[order.status.value]
